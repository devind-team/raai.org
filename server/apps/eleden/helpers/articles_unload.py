"""Осуществление выгрузки публикаций в различные форматы."""

import os
import posixpath
import re
import zipfile
from datetime import datetime, date
from functools import reduce
from os.path import join
from typing import Union, List, Dict

from bibtexparser.bibdatabase import BibDatabase
from bibtexparser.bwriter import BibTexWriter
from django.conf import settings
from django.template import Context
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from devind_helpers.generators import DocumentGenerator

from apps.core.models import User
from apps.core.schema import UserType
from apps.eleden.models import Article


class ArticlesUnload:
    """Класс для выполнения выгрузки."""

    headers: List[Dict[str, Union[str, int]]] = [
        {'name': 'order', 'title': '№', 'width': 5},
        {'name': 'name', 'title': 'Название работы', 'width': 30},
        {'name': 'type', 'title': 'Вид работы (печат. или рукопис)', 'width': 10},
        {'name': 'result', 'title': 'Результат (издание, номер, год, стр.)', 'width': 40},
        {'name': 'workload', 'title': 'Объем работы', 'width': 30},
        {'name': 'authors', 'title': 'Соавторы', 'width': 60},
    ]

    def __init__(self, user: User):
        """Генерация."""
        self.host = 'https://eleden.sbmpei.ru/'
        self.user = user
        self.articles = Article.objects \
            .select_related('kind', 'index', ) \
            .prefetch_related('author_set') \
            .filter(author__user=user)
        self.path: str = join(settings.DOCUMENTS_DIR, f'articles_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}')
        self.temp_path: str = f'storage/temp/{user.id}/'

    def xlsx(self):
        """Генерация выгрузки публикаций excel."""
        wb: Workbook = Workbook()
        ws = wb.active
        ws.title = f'Выгрузка публикаций'
        for position, header in enumerate(self.headers):
            ws.cell(1, position + 1, header['title']).alignment = Alignment(vertical="center", horizontal="center",
                                                                            wrap_text=True)
            ws.cell(2, position + 1, position + 1).alignment = Alignment(vertical="center", horizontal="center",
                                                                         wrap_text=True)
            ws.column_dimensions[ws.cell(1, position + 1).column_letter].width = header['width']
            ws.row_dimensions[1].height = 60
        start_row_position: int = 3
        for article in self.articles:
            delta_srp: int = 1
            for position, header in enumerate(self.headers):
                if header['name'] == 'order':
                    ws.cell(start_row_position, position + 1, f'{start_row_position - 2}')
                elif header['name'] == 'name':
                    ws.cell(start_row_position, position + 1, f'{article.name}')
                elif header['name'] == 'type':
                    ws.cell(start_row_position, position + 1, f'{article.additional["type"]}')
                elif header['name'] == 'result':
                    result = (article.additional['edition'],
                              article.additional['volume'],
                              str(article.year),
                              article.additional['pages'],
                              )
                    ws.cell(
                        start_row_position,
                        position + 1,
                        ', '.join(result))
                elif header['name'] == 'workload':
                    ws.cell(start_row_position, position + 1, f'{article.workload}')
                elif header['name'] == 'authors':
                    authors = ', '.join(
                        [author.name for author in article.author_set.all() if author.user_id != self.user.id])
                    ws.cell(start_row_position, position + 1, f'{authors}')
                else:
                    ws.cell(start_row_position, position + 1, article[header['name']])

            for row_index in range(start_row_position, start_row_position + delta_srp):
                ws.cell(row_index, 1).border = Border(left=Side(border_style='thin', color="FF000000"))
                ws.cell(row_index, len(self.headers)).border = Border(right=Side(border_style='thin', color="FF000000"))

            for position, header in enumerate(self.headers):
                if position == len(self.headers) - 1:
                    ws.cell(start_row_position, position + 1).border = Border(
                        top=Side(border_style='thin', color="FF000000"),
                        right=Side(border_style='thin', color="FF000000")
                    )
                else:
                    ws.cell(start_row_position, position + 1).border = Border(
                        top=Side(border_style='thin', color="FF000000"))

                ws.cell(start_row_position + delta_srp, position + 1).border = Border(
                    top=Side(border_style='thin', color="FF000000"))
            start_row_position += delta_srp
        path_output = f'{self.path}.xlsx'
        wb.save(path_output)
        return posixpath.relpath(path_output, settings.BASE_DIR)

    def docx(self):
        """Генерация выгрузки word."""
        path_xml: str = join(settings.BASE_DIR, 'apps', 'eleden', 'templates', 'articles', 'articles.xml')
        path_docx: str = join(settings.BASE_DIR, 'apps', 'eleden', 'templates', 'articles', 'articles.docx')

        def prepare(a, c):
            if c.kind.id not in a:
                c.kind.articles = []
                a[c.kind.id] = c.kind
            c.authors_rep = ', '.join([author.name for author in c.author_set.all() if author.user_id != self.user.id])
            a[c.kind.id].articles.append(c)
            return a

        kinds = reduce(prepare, self.articles, {})

        context: Context = Context({'kinds': kinds.values()})
        dg = DocumentGenerator(context, path_xml, path_docx)
        return dg.generate_docx(self.path).path

    def bibtex(self):
        """Генерация выгрузки в bibtex."""
        db = BibDatabase()
        writer = BibTexWriter()
        for article in self.articles:
            all_authors = [re.sub(r'([а-яё])\s*([А-ЯЁ])', r'\1, \2', a.name) for a in article.author_set.all()]
            db.entries = [{
                'title': article.name,
                'year': str(article.year),
                'author': ' and '.join(all_authors),
                'ID': article.user.last_name + str(article.year),
                'ENTRYTYPE': 'article',
                **article.additional
            }]
            bib_path = f'{self.temp_path}{article.name} {date.today()}.bib'
            with open(bib_path, 'w') as f:
                f.write(writer.write(db))
        path_output = f'{self.path}.zip'
        bib_zip = zipfile.ZipFile(path_output, 'w')
        for folder, subfolder, files in os.walk(f'{self.temp_path}'):
            for file in files:
                if file.endswith('.bib'):
                    bib_zip.write(os.path.join(folder, file), file, compress_type=zipfile.ZIP_DEFLATED)
                    os.remove(os.path.join(f'{self.temp_path}', file))
        bib_zip.close()
        return posixpath.relpath(path_output, settings.BASE_DIR)