"""Модуль отображения имен"""
import json
from itertools import groupby
from typing import Any, Optional, Union, Callable, Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .data import TeacherData, CoupleData, TeamData
from .wrong import WrongKind, WrongCoupleItem, WrongTeamItem, WrongItemsWrapper


class NamesMap:
    """Отображение имен"""

    def __init__(self, **kwargs):
        self._teams: dict[str, str] = kwargs.get('teams', {})
        self._work_kinds: dict[str, str] = kwargs.get('work_kinds', {})
        self._periods: dict[str, str] = kwargs.get('periods', {})
        self._teachers: dict[str, str] = kwargs.get('teachers', {})
        self._disciplines: dict[tuple[str, str], str] = kwargs.get('disciplines', {})
        self._couples_data: list[CoupleData] = kwargs.get('couples_data', [])
        self._teams_data: list[TeamData] = kwargs.get('teams_data', [])
        self._deleted_couples_data_ids: list[int] = kwargs.get('deleted_couples_data_ids', [])
        self._deleted_teams_data_ids: list[int] = kwargs.get('deleted_teams_data_ids', [])

    @classmethod
    def save_empty_to_json(cls, wrong_items_wrappers: list[WrongItemsWrapper], path: str) -> None:
        """Сохранение пустого шаблона в json.

        :param wrong_items_wrappers: список неправильных элементов с путями к файлам
        :param path: путь к файлу
        """

        template = cls._get_template(wrong_items_wrappers)
        with open(path, 'w', encoding='utf8') as json_file:
            json.dump(template, json_file, ensure_ascii=False, indent=2)

    @classmethod
    def save_empty_to_xlsx(cls, wrong_items_wrappers: list[WrongItemsWrapper], path: str) -> None:
        """Сохранение пустого шаблона в xlsx.

        :param wrong_items_wrappers: список неправильных элементов с путями к файлам
        :param path: путь к файлу
        """

        template = cls._get_template(wrong_items_wrappers)
        wb = Workbook()
        for wb_key, wb_value in template.items():
            ws = wb.create_sheet(wb_key)
            items = template[wb_key]
            for i, k in enumerate(items[0].keys(), 1):
                ws.column_dimensions[get_column_letter(i)].width = 50
                cell = ws.cell(1, i, k)
                cell.style = cls._ExcelStyles.title_style
            for i, item in enumerate(items, 2):
                for j, value in enumerate(item.values(), 1):
                    if isinstance(value, list):
                        cell = ws.cell(i, j, json.dumps(value, ensure_ascii=False))
                    elif isinstance(value, str):
                        cell = ws.cell(i, j, value)
                    else:
                        cell = ws.cell(i, j, '')
                    cell.style = cls._ExcelStyles.main_style
                    if wb_key == 'Дисциплины':
                        ws.row_dimensions[i].height = 52.2
        if len(wb.sheetnames) > 1:
            wb.remove_sheet(wb.get_sheet_by_name(wb.sheetnames[0]))
        else:
            wb.active.title = 'Неправильные данные отсутствуют'
        wb.save(path)

    @classmethod
    def load_from_json(cls, path: str) -> 'NamesMap':
        """Загрузка из json.

        :param path: путь к файлу
        :return: отображение имен
        """

        with open(path, 'r', encoding='utf8') as json_file:
            template = json.load(json_file)
            return cls._from_template(template)

    @classmethod
    def load_from_xlsx(cls, path: str) -> 'NamesMap':
        """Загрузка из xlsx.

        :param path: путь к файлу
        :return: отображение имен
        """

        wb = openpyxl.load_workbook(path)
        dict_keys = [
            'Группы', 'Виды работ', 'Периоды обучения',
            'Преподаватели', 'Дисциплины', 'Данные пар',
            'Данные групп'
        ]
        id_list_keys = ['Удаленные данные пар', 'Удаленные данные групп']
        template = {
            **{k: cls._dicts_from_sheet(wb[k]) if k in wb else [] for k in dict_keys},
            **{k: cls._id_list_from_sheet(wb[k]) if k in wb else [] for k in id_list_keys}
        }
        return cls._from_template(template)

    def get_team_data(self, team_data: TeamData) -> TeamData:
        """Получение правильных данных группы.

        :param team_data: возможно неправильные данные группы
        :return: правильные данные группы
        """

        right_team_data: Optional[TeamData] = None
        for map_team_data in self._teams_data:
            if team_data == map_team_data:
                right_team_data = TeamData(**{
                    **team_data.__dict__,
                    **{k: v for k, v in map_team_data.__dict__.items() if v is not None}
                })
        if right_team_data is not None:
            return right_team_data
        return TeamData(**{
            **team_data.__dict__,
            'short_name': self._get_team_short_name(team_data)
        })

    def get_couple_data(self, couple_data: CoupleData) -> CoupleData:
        """Получение правильных данных пары.

        :param couple_data: возможно неправильные данные пары
        :return: правильные данные пары
        """

        right_couple_data: Optional[CoupleData] = None
        for map_couple_data in self._couples_data:
            if couple_data == map_couple_data:
                right_couple_data = CoupleData(**{
                    **couple_data.__dict__,
                    **{k: v for k, v in map_couple_data.__dict__.items() if v is not None}
                })
        if right_couple_data is not None:
            return right_couple_data
        return CoupleData(**{
            **couple_data.__dict__,
            'work_kind': self._get_work_kind(couple_data),
            'periods': self._get_periods(couple_data),
            'teachers': self._get_teacher_data(couple_data),
            'discipline_words': self._get_discipline_words(couple_data),
        })

    def is_couple_data_deleted(self, couple_data: CoupleData) -> bool:
        """Проверка являются ли данные пары удаленнными.

        :param couple_data: данные пары
        :return: результат проверки
        """

        return couple_data.id in self._deleted_couples_data_ids

    def is_team_data_deleted(self, team_data: TeamData) -> bool:
        """Проверка являются ли данные группы удаленными.

        :param team_data: данные группы
        :return: результаты проверки
        """

        return team_data.id in self._deleted_teams_data_ids

    class _ExcelStyles:
        """Стили для ячеек Excel"""

        size = Side(border_style='thin', color='000000')
        border = Border(top=size, right=size, bottom=size, left=size)
        alignment = Alignment(vertical='top', wrap_text=True)
        title_style = NamedStyle('title')
        title_style.font = Font(name='Times New Roman', bold=True, size=14)
        title_style.border = border
        title_style.alignment = alignment
        main_style = NamedStyle('main')
        main_style.font = Font(name='Times New Roman', size=12)
        main_style.border = border
        main_style.alignment = alignment

    def _get_team_short_name(self, team_data: TeamData) -> str:
        """Получение правильного сокращенного названия группы.

        :param team_data: данные группы
        :return: правильное сокращенное название группы
        """

        if team_data.short_name in self._teams:
            return self._teams[team_data.short_name]
        return team_data.short_name

    def _get_work_kind(self, couple_data: CoupleData) -> str:
        """Получение правильного вида работы.

        :param couple_data: данные пары
        :return: правильный вид работы
        """

        if couple_data.work_kind in self._work_kinds:
            return self._work_kinds[couple_data.work_kind]
        return couple_data.work_kind

    def _get_periods(self, couple_data: CoupleData) -> Union[list[str], Callable[[str], bool]]:
        """Получение правильных периодов.

        :param couple_data: данные пары
        :return: правильные периоды
        """

        if callable(couple_data.periods):
            return couple_data.periods
        periods: list[str] = []
        for period in periods:
            if period in self._periods:
                periods.append(self._periods[period])
            else:
                periods.append(period)
        return periods

    def _get_teacher_data(self, couple_data: CoupleData) -> list[TeacherData]:
        """Получение правильных данных преподавателей.

        :param couple_data: данные пары
        :return: правильные данные преподавателей
        """

        teachers: list[TeacherData] = []
        for teacher_data in couple_data.teachers:
            str_teacher_data = str(teacher_data)
            if str_teacher_data in self._teachers:
                teacher_values = self._teachers[str_teacher_data].split(' ')
                if len(teacher_values) == 2:
                    teachers.append(TeacherData(None, *teacher_values))
                else:
                    teachers.append(TeacherData(*teacher_values))
            else:
                teachers.append(teacher_data)
        return teachers

    def _get_discipline_words(self, couple_data: CoupleData) -> list[str]:
        """Получение правильных слов названия дисциплины.

        :param couple_data: данные пары
        :return: правильные слова названия дисциплины
        """

        team_name = couple_data.team_data.short_name
        discipline_name = ' '.join(couple_data.discipline_words)
        if (team_name, discipline_name) in self._disciplines:
            return self._disciplines[(team_name, discipline_name)].split(' ')
        return couple_data.discipline_words

    @staticmethod
    def _get_russian_wrong_kind(wrong_kind: WrongKind) -> str:
        """Получение русского названия типа неправильного элемента.

        :param wrong_kind: тип неправильного элемента
        :return: русское название
        """

        if wrong_kind == wrong_kind.UNDISCOVERED:
            return 'Запись не найдена'
        if wrong_kind == wrong_kind.MULTIPLE:
            return 'Найдено несколько записей'
        if wrong_kind == wrong_kind.WRONG_DATA:
            return 'Неправильная запись'

    @staticmethod
    def _get_wrong_items(
            wrong_items_wrappers: list[WrongItemsWrapper],
            attr: str) -> Iterable[Union[WrongCoupleItem, WrongTeamItem]]:
        """Получение неправильных элементов по атрибуту.

        :param wrong_items_wrappers: список неправильных элементов с путями к файлам
        :param attr: атрибут
        :return: неправильные элементы
        """

        for wrong_items_wrappers in wrong_items_wrappers:
            for wrong_item in getattr(wrong_items_wrappers.wrong_items, attr):
                yield wrong_item

    @classmethod
    def _get_template_values(cls, wrong_items_wrappers: list[WrongItemsWrapper], attr: str) -> list[dict[str, str]]:
        """Получение значений для пустого шаблона.

        :param wrong_items_wrappers: список неправильных элементов с путями к файлам
        :param attr: атрибут неправильного элемента
        :return: значения для пустого шаблона
        """

        result_dict: dict[tuple[str, WrongKind], dict[str, str]] = {}
        for wrong_item in cls._get_wrong_items(wrong_items_wrappers, attr):
            if (wrong_item.value, wrong_item.kind) not in result_dict:
                result_dict[(wrong_item.value, wrong_item.kind)] = {
                    'Тип': cls._get_russian_wrong_kind(wrong_item.kind),
                    'Значение': wrong_item.value,
                    'Возможные значения': wrong_item.possible_values,
                    'Сообщение': wrong_item.message,
                    'Правильное значение': None
                }
        return list(result_dict.values())

    @classmethod
    def _get_template_disciplines(cls,
                                  wrong_items_wrappers: list[WrongItemsWrapper]) -> list[dict[str, Union[str, list]]]:
        """Получение дисциплин для пустого шаблона.

        :param wrong_items_wrappers: список неправильных элементов с путями к файлам
        :return: дисциплины для пустого шаблона
        """

        def key_func(wrong_couple: WrongCoupleItem) -> str:
            return wrong_couple.value + str(wrong_couple.kind) + str(wrong_couple.possible_values)

        sorted_disciplines = sorted(cls._get_wrong_items(wrong_items_wrappers, 'disciplines'), key=key_func)
        all_disciplines: list[dict[str, Union[str, list]]] = []
        for _, group in groupby(sorted_disciplines, key_func):
            group_list = list(group)
            all_disciplines.append({
                'Группы': [],
                'Тип': cls._get_russian_wrong_kind(group_list[0].kind),
                'Значение': group_list[0].value,
                'Возможные значения': group_list[0].possible_values,
                'Правильное значение': None
            })
            for wrong_item in group_list:
                if wrong_item.couple_data.team_data.short_name not in all_disciplines[-1]['Группы']:
                    all_disciplines[-1]['Группы'].append(wrong_item.couple_data.team_data.short_name)
        return all_disciplines

    @classmethod
    def _get_template(cls, wrong_items_wrappers: list[WrongItemsWrapper]) -> dict[str, list[dict]]:
        """Получение шаблона.

        :param wrong_items_wrappers: список неправильных элементов с путями к файлам
        :return: шаблон
        """

        template = {
            'Группы': cls._get_template_values(wrong_items_wrappers, 'teams'),
            'Виды работ': cls._get_template_values(wrong_items_wrappers, 'work_kinds'),
            'Периоды обучения': cls._get_template_values(wrong_items_wrappers, 'periods'),
            'Преподаватели': cls._get_template_values(wrong_items_wrappers, 'teachers'),
            'Дисциплины': cls._get_template_disciplines(wrong_items_wrappers)
        }
        for k, v in list(template.items()):
            if not len(v):
                del template[k]
        return template

    @staticmethod
    def _map_from_template_values(template_values: list[dict[str, str]]) -> dict[str, str]:
        """Получение отображения из значений шаблона.

        :param template_values: значения шаблона
        :return: отображение
        """

        result: dict[str, str] = {}
        for value in template_values:
            if value['Правильное значение']:
                result[value['Значение']] = value['Правильное значение']
        return result

    @staticmethod
    def _map_from_template_disciplines(template_disciplines: list[dict[str, str]]) -> dict[tuple[str, str], str]:
        """Получение отображения дисциплин из дисциплин шаблона.

        :param template_disciplines: дисциплины шаблона
        :return: отображение
        """

        result: dict[tuple[str, str], str] = {}
        for discipline in template_disciplines:
            if discipline['Правильное значение']:
                name = discipline['Значение']
                right_name = discipline['Правильное значение']
                for team in discipline['Группы']:
                    result[(team, name)] = right_name
        return result

    @staticmethod
    def _dicts_from_sheet(sheet: Worksheet) -> list[dict[str, Any]]:
        """Получение словарей из рабочего листа.

        :param sheet: рабочий лист
        :return: словарь
        """

        keys: list[str] = []
        for j in range(1, sheet.max_column + 1):
            keys.append(sheet.cell(1, j).value)
        result: list[dict[str, Any]] = []
        for row_cells in sheet.iter_rows(min_row=2):
            result.append({})
            for j, cell in enumerate(row_cells):
                try:
                    result[-1][keys[j]] = json.loads(cell.value)
                except (ValueError, TypeError):
                    result[-1][keys[j]] = cell.value
        return result

    @staticmethod
    def _id_list_from_sheet(sheet: Worksheet) -> list[int]:
        """Получение списка идентификаторов из рабочего листа.

        :param sheet: рабочий лист
        :return: список идентификаторов
        """

        result: list[int] = []
        for row_cells in sheet.rows:
            for cell in row_cells:
                if cell.value:
                    result.append(int(cell.value))
        return result

    @staticmethod
    def _create_couples_data(raw_couples_data: list[dict[str, Any]]) -> list[CoupleData]:
        """Создание данных пар из сырых данных пар.

        :param raw_couples_data: сырые данные пар
        :return: данные пар
        """

        couples_data: list[CoupleData] = []
        for raw in raw_couples_data:
            couples_data.append(CoupleData(
                work_kind=raw['Вид работы'],
                week_day=raw['День недели'],
                subgroup_number=raw.get('Номер подгруппы', None),
                discipline_words=raw['Слова названия дисциплины'],
                periods=raw.get('Периоды обучения', []),
                teachers=[TeacherData(**{
                    'position': teacher['Должность'],
                    'last_name': teacher['Фамилия'],
                    'initials': teacher['Инициалы']
                }) for teacher in raw.get('Преподаватели', [])],
                classrooms=raw['Аудитории'],
                class_numbers=raw['Номера занятия'],
                distance_learning=raw.get('Дистанционное обучение', False),
                id=raw['Идентификатор']
            ))
        return couples_data

    @classmethod
    def _create_teams_data(cls, raw_teams_data: list[dict]) -> list[TeamData]:
        """Создание данных групп из сырых данных групп.

        :param raw_teams_data: сырые данные групп
        :return: данные групп
        """

        teams_data: list[TeamData] = []
        for raw in raw_teams_data:
            teams_data.append(TeamData(
                short_name=raw['Сокращенное название'],
                course_number=raw['Номер курса'],
                semester_number=raw['Номер семестра'],
                couples=cls._create_couples_data(raw.get('Пары', [])),
                id=raw['Идентификатор']
            ))
        return teams_data

    @classmethod
    def _from_template(cls, template: dict[str, list[dict[str, str]]]) -> 'NamesMap':
        """Получение отображения имен из шаблона.

        :param template: шаблон
        :return: отображение имен
        """

        return cls(
            teams=cls._map_from_template_values(template.get('Группы', [])),
            work_kinds=cls._map_from_template_values(template.get('Виды работ', [])),
            periods=cls._map_from_template_values(template.get('Периоды обучения', [])),
            teachers=cls._map_from_template_values(template.get('Преподаватели', [])),
            disciplines=cls._map_from_template_disciplines(template.get('Дисциплины', [])),
            couples_data=cls._create_couples_data(template.get('Данные пар', [])),
            teams_data=cls._create_teams_data(template.get('Данные групп', [])),
            deleted_couples_data_ids=template.get('Удаленные данные пар', []),
            deleted_teams_data_ids=template.get('Удаленные данные групп', [])
        )
