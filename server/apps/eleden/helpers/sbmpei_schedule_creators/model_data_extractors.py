"""Модуль с извлекателями данных"""
import re
from abc import ABC, abstractmethod
from functools import reduce
from itertools import groupby
from typing import Any, Optional, Union, Iterable

import openpyxl
from django.core.files.uploadedfile import UploadedFile
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from .cache import Cache
from .couples_row_kind import get_couples_row_kind
from .couples_row_parser import CouplesRowParser, Result
from .data import ExcelInfo, TeamData, CoupleData, TeamModelData
from .model_data_factory import ModelDataFactory
from .names_map import NamesMap
from .wrong import WrongItems


class Extractor(ABC):
    """Базовый извлекатель данных моделей"""

    def __init__(self):
        self._teams_model_data: list[TeamModelData] = []
        self._wrong_items = WrongItems()
        self._cache = Cache()

    @abstractmethod
    def extract(self) -> None:
        """Извлечь данные моделей."""

        pass

    @property
    def teams_model_data(self) -> list[TeamModelData]:
        """Данные моделей групп"""

        return self._teams_model_data

    @property
    def wrong_items(self) -> WrongItems:
        """Неправильные элементы"""

        return self._wrong_items

    @property
    def cache(self) -> Cache:
        """Кеш"""

        return self._cache


class ExcelExtractor(Extractor):
    """Извлекатель данных моделей из Excel"""

    def __init__(self,
                 source: Union[str, UploadedFile],
                 semester_number: int,
                 names_map: Optional[NamesMap] = None,
                 teams_filter: Optional[dict[str, Any]] = None,
                 save_wrong_teams: bool = False):
        super().__init__()
        self._workbook = openpyxl.load_workbook(source)
        self._semester_number = semester_number
        self._names_map: NamesMap = names_map if names_map else NamesMap()
        self._teams_filter: dict[str, Any] = teams_filter if teams_filter else {}
        self._save_wrong_teams = save_wrong_teams
        self.extract()

    _days = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота']

    def extract(self) -> None:
        for sheet in self._workbook.worksheets:
            self._parse_sheet(sheet)

    def _parse_sheet(self, sheet: Worksheet) -> None:
        """Парсинг листа.

        :param sheet: рабочий лист книги
        """

        for cell, column, course_number in self._get_team_columns(sheet):
            team_data = TeamData(ExcelInfo(sheet, [cell]), cell.value, course_number, self._semester_number)
            for day, start_row, end_row in self._get_days_rows(sheet):
                couples_data: list[CoupleData] = self._parse_day(sheet, day, start_row, end_row, column)
                for couple_data in couples_data:
                    couple_data.team_data = team_data
                    team_data.couples.append(couple_data)
            self._create_team_model_data(team_data)

    def _parse_day(self,
                   sheet: Worksheet,
                   week_day: int,
                   start_row: int,
                   end_row: int,
                   column: int) -> list[CoupleData]:
        """Парсинг дня.

        :param sheet: рабочий лист книги
        :param week_day: день недели
        :param start_row: строка начала дня
        :param end_row: строка окончания дня
        :param column: столбец дня
        :return: список данных пар
        """

        result: list[CoupleData] = []
        previous_parser_results: list[Result] = []
        class_number = 1
        for row in range(start_row, end_row, 2):
            excel_info = self._get_excel_info(sheet, row, column)
            couples_row_kind = get_couples_row_kind(sheet, excel_info.cells)
            previous_parser_results = CouplesRowParser(
                excel_info,
                couples_row_kind,
                week_day,
                class_number,
                previous_parser_results
            ).results
            for parser_result in (res for res in previous_parser_results if res.couples_data):
                for couple_data in parser_result.couples_data:
                    result.append(couple_data)
            class_number += 1
        return result

    def _create_team_model_data(self, team_data: TeamData) -> None:
        """Создание данных модели группы.

        :param team_data: данные группы
        """

        wrong_teams_data, teams_model_data, wrongs = ModelDataFactory(
            [team_data],
            self._teams_filter,
            self._names_map,
        ).creation_results
        if len(teams_model_data):
            self._cache = self._cache | Cache(wrong_teams_data)
            self._teams_model_data.extend(teams_model_data)
            for wrong in wrongs:
                self._wrong_items.add_from_wrong_dict(wrong)
        elif self._save_wrong_teams:
            self._cache = self._cache | Cache(wrong_teams_data)
            for wrong in wrongs:
                self._wrong_items.add_from_wrong_dict(wrong)

    def _get_team_columns(self, sheet: Worksheet) -> Iterable[tuple[Cell, int, int]]:
        """Получение столбцов групп.

        :param sheet: рабочий лист книги
        :return: (ячейка группы, номер столбца, номер курса)
        """

        course_numbers = sorted(map(int, re.findall(r'\d+', sheet.title)))
        row = 2
        column = 3
        cell = sheet.cell(row, column)
        all_cells_with_columns: list[tuple[Cell, int]] = []
        while cell.value is not None:
            all_cells_with_columns.append((cell, column))
            column += 3
            cell = sheet.cell(row, column)
        for i, cells_with_columns in enumerate(self._group_by_team_number(all_cells_with_columns)):
            for cell_with_column in cells_with_columns:
                yield *cell_with_column, course_numbers[0]

    @classmethod
    def _get_days_rows(cls, sheet: Worksheet) -> Iterable[tuple[int, int, int]]:
        """Получение начальных и конечных строк дней рабочего листа.

        :param sheet: рабочий лист книги
        :return: (номер дня, начальная строка, конечная строка)
        """

        for i, day in enumerate(cls._days, 1):
            day_rows = cls._get_day_rows(sheet, day)
            if day_rows is not None:
                yield i, *day_rows

    @staticmethod
    def _get_day_rows(sheet: Worksheet, day: str) -> Optional[tuple[int, int]]:
        """Получение начальной и конечной строки дня.

        :param sheet: рабочий лист книги
        :param day: день недели
        :return: (начальная строка, конечная строка)
        """

        row = 3
        column = 1
        cell = sheet.cell(row, column)
        while cell.value != day:
            if cell.value is None and \
                    not reduce(
                        lambda acc, curr: acc or isinstance(sheet.cell(row + curr, column), MergedCell), [1, 2, 3],
                        False):
                return None
            row += 1
            cell = sheet.cell(row, column)
        start_row = row
        row += 1
        cell = sheet.cell(row, column)
        while isinstance(cell, MergedCell):
            row += 1
            cell = sheet.cell(row, column)
        end_row = row - 1
        return start_row, end_row

    @staticmethod
    def _group_by_team_number(cells_with_columns: list[tuple[Cell, int]]) -> Iterable[tuple[Cell, int]]:
        """Группировка по номеру группы по убыванию.

        :param cells_with_columns: ячейки с номерами столбцов
        :return: ячейки с померами столбцов сгруппированные по номеру группы по убыванию
        """

        def key_func(cell_with_column: tuple[Cell, int]) -> int:
            return int(re.findall(r'\d+', cell_with_column[0].value)[0])

        for _, items in groupby(sorted(cells_with_columns, key=key_func, reverse=True), key_func):
            yield items

    @staticmethod
    def _get_excel_info(sheet: Worksheet, row: int, column: int) -> ExcelInfo:
        """Получение информации о положении в книге.

        :param sheet: рабочий лист книги
        :param row: строка первой ячейки
        :param column: столбец первой ячейки
        :return: информация о потожении в книге.
        """

        return ExcelInfo(
            sheet,
            [sheet.cell(row, column), sheet.cell(row, column + 1), sheet.cell(row, column + 2),
             sheet.cell(row + 1, column), sheet.cell(row + 1, column + 1), sheet.cell(row + 1, column + 2)]
        )


class CacheExtractor(Extractor):
    """Извлекатель данных моделей из кеша"""

    def __init__(self,
                 source_cache: Cache,
                 names_map: Optional[NamesMap] = None,
                 teams_filter: Optional[dict[str, Any]] = None):
        super().__init__()
        self._source_cache = source_cache
        self._names_map: NamesMap = names_map if names_map else NamesMap()
        self._teams_filter: dict[str, Any] = teams_filter if teams_filter else {}
        self.extract()

    def extract(self) -> None:
        wrong_teams_data, teams_model_data, wrongs = ModelDataFactory(
            self._source_cache.teams,
            self._teams_filter,
            self._names_map,
        ).creation_results
        self._teams_model_data = teams_model_data
        self._cache = Cache(wrong_teams_data)
        for wrong in wrongs:
            self._wrong_items.add_from_wrong_dict(wrong)
