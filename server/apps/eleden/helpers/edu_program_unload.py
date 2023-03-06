import posixpath
from datetime import datetime
from functools import reduce
from os.path import join
from typing import List, Dict, Union

from django.conf import settings
from django.db.models import Prefetch
from django.template import Template, Context
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

from apps.eleden.models import Discipline, MethodologicalSupport


class EduProgramUnload:

    headers: List[Dict[str, Union[str, int]]] = [
        {'name': 'direction__code', 'title': 'Код', 'width': 10},
        {'name': 'direction__name', 'title': 'Наименование профессии, специальности, направления подготовки', 'width': 30},
        {'name': 'direction__edu_service__name', 'title': 'Уровень образования', 'width': 15},
        {'name': 'name', 'title': 'Образовательная программа, направленность, профиль, шифр и наименование научной специальности', 'width': 30},
        {'name': 'edu_form__name', 'title': 'Форма обучения', 'width': 15},
        {'name': 'syllabus', 'title': 'Учебный план', 'width': 15},
        {'name': 'calendar', 'title': 'Календарный учебный график', 'width': 15},
        {'name': 'description', 'title': 'Описание образовательной программы. РПД образовательной программы.', 'width': 100},
        {'name': 'annotations', 'title': 'Аннотации', 'width': 80},
        {'name': 'methodological_support', 'title': 'Методические и иные документы, разработанные образовательной организацией для обеспечения образовательного процесса', 'width': 50}
    ]

    def __init__(self, edu_programs: list, host: str, date_sign: str):
        """Генерация."""
        self.host = host
        self.date_sign = date_sign
        self.edu_programs = edu_programs
        self.path: str = join(settings.DOCUMENTS_DIR, f'edu_programs_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}')
        disciplines = Discipline.objects \
            .prefetch_related(Prefetch('methodologicalsupport_set')) \
            .filter(edu_program__in=[ep['id'] for ep in self.edu_programs]) \
            .exclude(pk__in=Discipline.objects.filter(parent_id__isnull=False).values_list('parent_id', flat=True)) \
            .all()
        disciplines = sorted(disciplines, key=lambda discipline: discipline.order)
        for ep in self.edu_programs:
            ep['disciplines'] = [d for d in disciplines if d.edu_program_id == ep['id']]
            ep['methodological_support'] = reduce(lambda a, c: [*a, *c.methodologicalsupport_set.all()], ep['disciplines'], [])

    def html(self):
        """Генерация выгрузки html."""
        path_template: str = join(settings.BASE_DIR, 'apps', 'eleden', 'templates', 'edu_programs', 'index.html')
        with open(path_template) as f:
            template: Template = Template(f.read())
        context: Context = Context({'edu_programs': self.edu_programs, 'host': self.host, 'date_sign': self.date_sign})
        path_output = f'{self.path}.html'
        with open(path_output, 'w+') as f:
            f.write(template.render(context))
        return posixpath.relpath(path_output, settings.BASE_DIR)

    def excel(self):
        """Генерация выгрузки excel."""
        wb: Workbook = Workbook()
        ws = wb.active
        ws.title = f'Выгрузка ООП'
        # Формируем заголовок
        for position, header in enumerate(self.headers):
            ws.cell(1, position + 1, header['title']).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(2, position + 1, position + 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[ws.cell(1, position + 1).column_letter].width = header['width']
            ws.row_dimensions[1].height = 60
        start_row_position: int = 3
        for edu_program in self.edu_programs:
            delta_srp: int = 1  # Образовательная программа
            ep_merge: Dict[int, int] = {}
            for d in edu_program['disciplines']:
                count_ms = sum([1 for ms in edu_program['methodological_support'] if ms.discipline_id == d.id])
                ep_merge[d.id] = count_ms if count_ms > 0 else 1
                delta_srp += ep_merge[d.id]
            for position, header in enumerate(self.headers):
                if header['name'] == 'name':
                    ws.cell(start_row_position, position + 1, f'{edu_program["name"]} ({edu_program["admission"]})').alignment = Alignment(vertical="center", wrap_text=True)
                elif header['name'] == 'description':
                    if edu_program['description']:
                        f: str = f'=HYPERLINK("{self.host}{edu_program["description"]}", "Описание образовательной программы")'
                        ws.cell(start_row_position, position + 1, f).alignment = Alignment(vertical="center", wrap_text=True)
                    self.__write_discipline_to_cell(ws, ep_merge, position, start_row_position, edu_program['disciplines'], 'work_program')
                elif header['name'] in ['syllabus', 'calendar']:
                    if edu_program[header['name']]:
                        f: str = f'=HYPERLINK("{self.host}{edu_program[header["name"]]}", "{header["title"]}")'
                        ws.cell(start_row_position, position + 1, f).alignment = Alignment(vertical="center", wrap_text=True)
                elif header['name'] == 'annotations':
                    self.__write_discipline_to_cell(ws, ep_merge, position, start_row_position, edu_program['disciplines'], 'annotation')
                elif header['name'] == 'methodological_support':
                    self.__write_methodological_support_to_cell(ws, position, start_row_position, edu_program['disciplines'], edu_program['methodological_support'])
                else:
                    ws.cell(start_row_position, position + 1, edu_program[header['name']]).alignment = Alignment(vertical="center", wrap_text=True)

            for row_index in range(start_row_position, start_row_position + delta_srp):
                ws.cell(row_index, 1).border = Border(left=Side(border_style='thin', color="FF000000"))
                ws.cell(row_index, len(self.headers)).border = Border(right=Side(border_style='thin', color="FF000000"))

            for position, header in enumerate(self.headers):
                if position == len(self.headers) - 1:
                    ws.cell(start_row_position, position + 1).border = Border(
                        top=Side(border_style='thin', color="FF000000"),
                        right=Side(border_style='thin', color="FF000000")
                    )
                else:
                    ws.cell(start_row_position, position + 1).border = Border(top=Side(border_style='thin', color="FF000000"))

                ws.cell(start_row_position + delta_srp, position + 1).border = Border(bottom=Side(border_style='thin', color="FF000000"))

                if header['name'] in ['direction__code', 'direction__name', 'name', 'direction__edu_service__name', 'edu_form__name', 'syllabus', 'calendar']:
                    ws.merge_cells(
                        start_row=start_row_position,
                        start_column=position + 1,
                        end_row=start_row_position + delta_srp - 1,
                        end_column=position + 1
                    )
            start_row_position += delta_srp
        path_output = f'{self.path}.xlsx'
        wb.save(path_output)
        return posixpath.relpath(path_output, settings.BASE_DIR)

    def __write_discipline_to_cell(
            self,
            ws,
            ep_merge: Dict[int, int],
            position: int,
            srp: int,
            items: List[Discipline],
            fn: str
    ):
        """Пишем дисциплину в ячейку."""
        for d in items:
            if getattr(d, fn):
                value: str = f'=HYPERLINK("{self.host}{getattr(d, fn)}", "{d.code} {d.name}")'
                font: Font = Font(bold=d.kind_id == 2)
            else:
                value: str = f"{d.code} {d.name}"
                font: Font = Font(bold=d.kind_id == 2, color='00FF0000')
            srp += 1
            ws.cell(srp, position + 1, value).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(srp, position + 1).font = font
            srp += ep_merge[d.id] - 1
            if ep_merge[d.id] > 1:
                ws.merge_cells(
                    start_row=srp - ep_merge[d.id] + 1,
                    start_column=position + 1,
                    end_row=srp,
                    end_column=position + 1
                )

    def __write_methodological_support_to_cell(self, ws, position: int, srp: int, items: List[Discipline], ms: List[MethodologicalSupport]):
        for d in items:
            methodological_supports = [m for m in ms if m.discipline_id == d.id]
            if len(methodological_supports) == 0:
                srp += 1
                continue
            for met_sub in methodological_supports:
                srp += 1
                value: str = f'=HYPERLINK("{self.host}{met_sub.src}", "{met_sub.name}")'
                ws.cell(srp, position + 1, value).alignment = Alignment(vertical="center", wrap_text=True)
